import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


def export_to_excel(report_data: dict, config: dict):
    os.makedirs("data", exist_ok=True)
    filepath = "data/monthly_reports.xlsx"

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Reports"
        ws.append(["Month", "Year", "Field", "Value", "Exported At"])

    month   = config["report_month"]
    year    = config["report_year"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for field, value in report_data.items():
        ws.append([month, year, field, value, now_str])

    wb.save(filepath)
    print(f"Excel saved — {len(report_data)} rows appended to {filepath}")
